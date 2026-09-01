# -*- coding: utf-8 -*-
"""Diferença estrutural: o que eu entreguei x o que existe agora."""
import openpyxl, difflib, json
from openpyxl.utils import get_column_letter

ATUAL = "/root/.claude/uploads/41fbd151-f5e2-530d-88f9-9b26c68b08f9/51670d45-Proposta_Reordenamento_2027_SEDUC__PI_1.xlsx"
wm = openpyxl.load_workbook("Reordenamento_2027_v1_v2_v3.xlsx", read_only=False)
wa = openpyxl.load_workbook(ATUAL, read_only=False)

def heads(wb, aba):
    ws = wb[aba]
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        out.append(str(v).replace("\n", " ").strip() if v not in (None, "") else "")
    while out and out[-1] == "": out.pop()
    return out

print("ABAS")
for n in wa.sheetnames:
    print(f"  {'=' if n in wm.sheetnames else '+ NOVA':>6s}  {n}")
for n in wm.sheetnames:
    if n not in wa.sheetnames: print(f"  {'- SAIU':>6s}  {n}")

print("\nCOLUNAS — só as abas que existem nos dois")
mudou = {}
for aba in wa.sheetnames:
    if aba not in wm.sheetnames: continue
    a, b = heads(wm, aba), heads(wa, aba)
    if a == b: continue
    print(f"\n  ── {aba}   ({len(a)} → {len(b)} colunas)")
    sm = difflib.SequenceMatcher(None, a, b)
    reg = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal": continue
        if tag in ("delete", "replace"):
            for i in range(i1, i2):
                print(f"       - saiu   {get_column_letter(i+1):>3s}  {a[i][:56]}")
                reg.append(("saiu", a[i]))
        if tag in ("insert", "replace"):
            for j in range(j1, j2):
                print(f"       + entrou {get_column_letter(j+1):>3s}  {b[j][:56]}")
                reg.append(("entrou", b[j]))
    # coluna que apenas mudou de lugar
    movidas = [(x, a.index(x), b.index(x)) for x in a if x and x in b and a.index(x) != b.index(x)]
    for nome, i, j in movidas:
        print(f"       ↔ mudou  {get_column_letter(i+1)} → {get_column_letter(j+1)}  {nome[:52]}")
    mudou[aba] = {"antes": a, "agora": b}
json.dump(mudou, open("estrutura.json", "w"), ensure_ascii=False)
