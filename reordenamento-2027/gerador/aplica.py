# -*- coding: utf-8 -*-
"""
Aplica a base real da UETEP SOBRE o arquivo de vocês.
Regra: só escreve onde precisa; nada mais é tocado.
"""
import openpyxl, pickle, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
sys.path.insert(0, ".")
from motor import numero_
from agrega import mapa_gdi, processar_turmas
from uetep_motor import ler_oferta, salas_2027, inep_ as ip

ATUAL  = "/root/.claude/uploads/41fbd151-f5e2-530d-88f9-9b26c68b08f9/51670d45-Proposta_Reordenamento_2027_SEDUC__PI_1.xlsx"
OFERTA = "/root/.claude/uploads/41fbd151-f5e2-530d-88f9-9b26c68b08f9/8e7688c9-OFERTA_2027_BASE.xlsx"
ENTREGUE = "Reordenamento_2027_v1_v2_v3.xlsx"
SAIDA  = "Proposta_Reordenamento_2027_com_UETEP.xlsx"

ARIAL = "Arial"
BORDA = Border(*[Side(style="thin", color="D1D5DB")] * 4)
TEAL, AMBAR, ROXO = "0F766E", "FBBF24", "7C3AED"

def cabecalho(ws, c, texto, cor=TEAL, fg="FFFFFF"):
    x = ws.cell(1, c, texto)
    x.fill = PatternFill("solid", fgColor=cor)
    x.font = Font(name=ARIAL, size=8, bold=True, color=fg)
    x.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    x.border = BORDA

def celula(ws, r, c, v, centro=True, wrap=False, fundo=None, negrito=False):
    x = ws.cell(r, c, v)
    x.font = Font(name=ARIAL, size=9, bold=negrito)
    x.alignment = Alignment(vertical="top", wrap_text=wrap,
                            horizontal="center" if centro else "general")
    x.border = BORDA
    if fundo: x.fill = PatternFill("solid", fgColor=fundo)
    return x

# ══════════════════════════════════════════════════ entrada
wo = openpyxl.load_workbook(OFERTA, data_only=True)
OI = [list(r) for r in wo["OFERTA INTEGRAL 2027"].iter_rows(values_only=True)]
CO = [list(r) for r in wo["CONTINUIDADE 2027"].iter_rows(values_only=True)]
SU = [list(r) for r in wo["SUBSEQUENTE 2027"].iter_rows(values_only=True)]
oferta, descartadas = ler_oferta(OI, CO, SU)

wb = openpyxl.load_workbook(ATUAL, data_only=False)
wv = openpyxl.load_workbook(ATUAL, data_only=True)
we = openpyxl.load_workbook(ENTREGUE, data_only=True)

turmas = [list(r) for r in wv["Turmas"].iter_rows(values_only=True)]
gdi    = mapa_gdi([list(r) for r in wv["Base GDI"].iter_rows(values_only=True)])
matrizes, ordem_m, anexos, ordem_a = processar_turmas(turmas, gdi)

mudancas = []
def log(t): mudancas.append(t); print("  ·", t)

# ══════════════════════════════════════════════════ 1. abas da base real
print("\n1. Abas da base real da UETEP")
for nome_novo, dados, larg in [
    ("UETEP · Oferta Integral 2027", OI, [8, 22, 12, 32, 10, 46, 12, 18, 14]),
    ("UETEP · Continuidade 2027",    CO, [8, 22, 12, 32, 46, 12, 11]),
    ("UETEP · Subsequente 2027",     SU, [8, 22, 12, 32, 12, 46, 26, 10, 26]),
]:
    if nome_novo in wb.sheetnames: del wb[nome_novo]
    ws = wb.create_sheet(nome_novo)
    linhas = [r for r in dados[1:] if ip(r[2]) and ip(r[2]) != "None"]
    for c, h in enumerate(dados[0], start=1):
        cabecalho(ws, c, str(h or "").replace("\n", " "))
    for r, linha in enumerate(linhas, start=2):
        for c, v in enumerate(linha, start=1):
            celula(ws, r, c, v, centro=(c not in (2, 4, 6)), wrap=(c in (4, 6)))
    for c, w in enumerate(larg, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(dados[0])), len(linhas) + 1)
    log("aba '%s' criada — %d linhas (%d de total descartada)"
        % (nome_novo, len(linhas), len(dados) - 1 - len(linhas)))

# ══════════════════════════════════════════════════ 2. Base UETEP refeita
print("\n2. Base UETEP")
if "Base UETEP" in wb.sheetnames:
    wb["Base UETEP"].title = "antiga base uetep"
    log("aba 'Base UETEP' antiga preservada como 'antiga base uetep'")

ws = wb.create_sheet("Base UETEP")
H = ["INEP", "GRE", "Município", "Escola",
     "1ª série 2027 · turmas", "1ª série 2027 · cursos", "1ª série 2027 · alunos previstos",
     "2ª série 2027 · turmas", "2ª série 2027 · alunos", "2ª série 2027 · cursos",
     "3ª série 2027 · turmas", "3ª série 2027 · alunos", "3ª série 2027 · cursos",
     "Subsequente 2027 · turmas", "Subsequente 2027 · alunos", "Subsequente 2027 · cursos",
     "Total de turmas 2027", "Salas necessárias 2027 (oferta real)", "Composição das salas"]
for c, h in enumerate(H, start=1): cabecalho(ws, c, h)
ordem = sorted(oferta, key=lambda i: (oferta[i]["gre"], oferta[i]["municipio"], oferta[i]["escola"]))
for r, i in enumerate(ordem, start=2):
    o = oferta[i]
    u = matrizes.get(i)
    s = salas_2027(u, o) if u else {"salas": "", "integral": "", "manha": "", "tarde": "", "noite": ""}
    comp = ("" if not u else "integral %d + máx(manhã %d, tarde %d) + noite %d"
            % (s["integral"], s["manha"], s["tarde"], s["noite"]))
    vals = [int(i), o["gre"], o["municipio"], o["escola"],
            int(o["of1_turmas"]), o["of1_detalhe"], int(o["of1_alunos"]),
            int(o["co2_turmas"]), int(o["co2_alunos"]), o["co2_detalhe"],
            int(o["co3_turmas"]), int(o["co3_alunos"]), o["co3_detalhe"],
            int(o["sub_turmas"]), int(o["sub_alunos"]), o["sub_detalhe"],
            int(o["total_turmas"]), s["salas"], comp]
    for c, v in enumerate(vals, start=1):
        celula(ws, r, c, v, centro=(c not in (3, 4, 6, 10, 13, 16, 19)),
               wrap=(c in (6, 10, 13, 16, 19)))
for c, w in enumerate([11, 14, 20, 32, 12, 44, 14, 12, 12, 44, 12, 12, 44, 12, 12, 32, 12, 14, 34], start=1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 44
ws.freeze_panes = "E2"
ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(H)), len(ordem) + 1)
log("aba 'Base UETEP' refeita a partir da oferta real — %d escolas" % len(ordem))

pickle.dump({"oferta": oferta, "ordem": ordem}, open("oferta.pkl", "wb"))
wb.save(SAIDA)
print("\nparcial salvo:", SAIDA)

# ══════════════════════════════════════════════════ 3. Base Tratada
print("\n3. Base Tratada — colunas da oferta real de 2027")
bt = wb["Base Tratada"]; btv = wv["Base Tratada"]
BASE = 30                                   # colunas que já existiam
NOVAS_BT = ["Oferta 2027 · 2ª série (turmas)", "Oferta 2027 · 2ª série (alunos)",
            "Oferta 2027 · 3ª série (turmas)", "Oferta 2027 · 3ª série (alunos)",
            "Oferta 2027 · subsequente (turmas)", "Oferta 2027 · subsequente (alunos)",
            "Oferta 2027 · 2ª e 3ª série (cursos)", "Salas 2027 · composição",
            "Oferta 2027 · tem oferta?"]
# colunas que já existiam e mudam de fonte
RENOMEIA_BT = {19: "Oferta 2027 · 1ª série (alunos previstos)",
               20: "Oferta 2027 · 1ª série (turmas)",
               21: "Oferta 2027 · 1ª série (cursos)",
               28: "Salas necessárias 2027 (oferta real)"}
for c, h in RENOMEIA_BT.items(): cabecalho(bt, c, h)
for j, h in enumerate(NOVAS_BT): cabecalho(bt, BASE + 1 + j, h)

n_bt = 0
for r in range(2, btv.max_row + 1):
    i = ip(btv.cell(r, 1).value)
    if not i or i == "None": continue
    o = oferta.get(i)
    u = matrizes.get(i)
    sl = salas_2027(u, o) if u else {"salas": "", "integral": 0, "manha": 0, "tarde": 0, "noite": 0}
    comp = ("integral %d + máx(manhã %d, tarde %d) + noite %d"
            % (sl["integral"], sl["manha"], sl["tarde"], sl["noite"])) if u else "—"
    # colunas antigas que mudaram de fonte
    celula(bt, r, 19, int(o["of1_alunos"]) if o else 0)
    celula(bt, r, 20, int(o["of1_turmas"]) if o else 0)
    celula(bt, r, 21, o["of1_detalhe"] if o else "— sem oferta de EM em 2027",
           centro=False, wrap=True)
    celula(bt, r, 28, sl["salas"])
    vals = ([int(o["co2_turmas"]), int(o["co2_alunos"]),
             int(o["co3_turmas"]), int(o["co3_alunos"]),
             int(o["sub_turmas"]), int(o["sub_alunos"]),
             o["co2_detalhe"] + " || " + o["co3_detalhe"], comp, "SIM"] if o
            else [0, 0, 0, 0, 0, 0, "—", comp, "NÃO"])
    for j, v in enumerate(vals):
        celula(bt, r, BASE + 1 + j, v, centro=(j not in (6, 7)), wrap=(j in (6, 7)))
    n_bt += 1
for j, w in enumerate([12, 12, 12, 12, 12, 12, 46, 34, 11]):
    bt.column_dimensions[get_column_letter(BASE + 1 + j)].width = w
bt.auto_filter.ref = "A1:%s%d" % (get_column_letter(BASE + len(NOVAS_BT)), btv.max_row)
log("Base Tratada: %d colunas novas em %d linhas (colunas A..AD intactas)" % (len(NOVAS_BT), n_bt))

# ══════════════════════════════════════════════════ 4. re-semeadura das ★
print("\n4. Colunas ★ — semeadura da oferta real, preservando o que foi digitado")

def valores_entregues(aba, col):
    """O que EU entreguei naquela célula — serve para saber o que ainda é semente."""
    ws = we[aba]
    out = {}
    for r in range(2, ws.max_row + 1):
        k = ip(ws.cell(r, 1).value)
        if k and k != "None": out[k] = ws.cell(r, col).value
    return out

def n(x):
    try: return round(float(x), 4)
    except (TypeError, ValueError): return None

preservadas, resemeadas, sem_oferta = [], 0, 0

def resemear(aba, col, campo, rotulo):
    """Troca a semente antiga pela oferta real; não encosta em célula editada."""
    global resemeadas
    ws, wsv = wb[aba], wv[aba]
    antigo = valores_entregues(aba, col)
    trocadas = 0
    for r in range(2, wsv.max_row + 1):
        i = ip(wsv.cell(r, 1).value)
        if not i or i == "None": continue
        atual = wsv.cell(r, col).value
        if atual in (None, ""): continue
        if n(atual) != n(antigo.get(i)):
            preservadas.append((aba, rotulo, i, atual))
            continue
        o = oferta.get(i)
        novo = int(o[campo]) if o else 0
        if n(novo) != n(atual):
            ws.cell(r, col, novo)
            trocadas += 1
    resemeadas += trocadas
    log("%s · %s: %d célula(s) atualizadas pela oferta real" % (aba[-2:], rotulo, trocadas))

# V2 (38 colunas): M = ★1ª série · X = ★salas
resemear("Reordenamento 2027 V2", 13, "of1_turmas", "★ 1ª SÉRIE 2027")
# V3 (48 colunas): O = ★1ª · S = ★2ª · T = ★3ª
resemear("Reordenamento 2027 V3", 15, "of1_turmas", "★ 1ª SÉRIE 2027")
resemear("Reordenamento 2027 V3", 19, "co2_turmas", "★ 2ª SÉRIE 2027")
resemear("Reordenamento 2027 V3", 20, "co3_turmas", "★ 3ª SÉRIE 2027")

# salas: recalculadas com a oferta real
def resemear_salas(aba, col, rotulo):
    global resemeadas
    ws, wsv = wb[aba], wv[aba]
    antigo = valores_entregues(aba, col)
    trocadas = 0
    for r in range(2, wsv.max_row + 1):
        i = ip(wsv.cell(r, 1).value)
        if not i or i == "None": continue
        atual = wsv.cell(r, col).value
        if atual in (None, ""): continue
        if n(atual) != n(antigo.get(i)):
            preservadas.append((aba, rotulo, i, atual)); continue
        u = matrizes.get(i)
        if not u: continue
        novo = salas_2027(u, oferta.get(i))["salas"]
        if n(novo) != n(atual):
            ws.cell(r, col, novo); trocadas += 1
    resemeadas += trocadas
    log("%s · %s: %d célula(s) recalculadas com a oferta real" % (aba[-2:], rotulo, trocadas))

resemear_salas("Reordenamento 2027 V2", 24, "★ SALAS NECESSÁRIAS")
resemear_salas("Reordenamento 2027 V3", 31, "★ SALAS NECESSÁRIAS")

print("\n  células preservadas por terem sido editadas por vocês:", len(preservadas))
for x in preservadas[:10]: print("     ", x)
pickle.dump(preservadas, open("preservadas.pkl", "wb"))
wb.save(SAIDA)
print("\nparcial salvo:", SAIDA)

# ══════════════════════════════════════════════════ 5. fórmulas da V2 e V3
print("\n5. Fórmulas apontando para a oferta real")
NBT = 577                                    # última linha da Base Tratada
def bt_(chave, col, err='""'):
    return "IFERROR(VLOOKUP(%s,'Base Tratada'!$A$2:$AM$%d,%d,FALSE),%s)" % (chave, NBT, col, err)

# ---- V3: colunas U, V, W deixam de ser a UETEP inventada
v3, v3v = wb["Reordenamento 2027 V3"], wv["Reordenamento 2027 V3"]
NL = 577
cabecalho(v3, 21, "Oferta 2027 ·\n1ª série (turmas)")
cabecalho(v3, 22, "Oferta 2027 ·\ncursos da 1ª série")
cabecalho(v3, 23, "Oferta 2027 ·\n2ª / 3ª / subseq.")
for r in range(2, NL + 1):
    A = "$A%d" % r
    v3.cell(r, 21, "=" + bt_(A, 20, "0"))
    v3.cell(r, 22, "=" + bt_(A, 21, '"—"'))
    v3.cell(r, 23, '=IF(%s="","",%s&" de 2ª | "&%s&" de 3ª | "&%s&" subseq.")'
            % (A, bt_(A, 31, "0"), bt_(A, 33, "0"), bt_(A, 35, "0")))
log("V3 · U/V/W repontadas para a oferta real (eram a UETEP por pré-matrícula)")

# ---- V3: AD = salas calculadas, agora sobre a oferta real
for r in range(2, NL + 1):
    A = "$A%d" % r
    v3.cell(r, 30, '=IF(%s="","",MAX(0,%s+($O%d-%s)+($S%d-%s)+($T%d-%s)))'
            % (A, bt_(A, 28, "0"), r, bt_(A, 20, "0"), r, bt_(A, 31, "0"), r, bt_(A, 33, "0")))
log("V3 · AD (salas calculadas) agora parte da oferta real de 2027")

# ---- V3: duas colunas novas no fim
cabecalho(v3, 49, "Oferta 2027 ·\nalunos reais 2ª/3ª", ROXO)
cabecalho(v3, 50, "★ x Oferta 2027 ·\ndivergência", ROXO)
for r in range(2, NL + 1):
    A = "$A%d" % r
    v3.cell(r, 49, '=IF(%s="","",%s+%s&" aluno(s)")' % (A, bt_(A, 32, "0"), bt_(A, 34, "0")))
    v3.cell(r, 50, ('=IF(%s="","",IF(AND($O%d=%s,$S%d=%s,$T%d=%s),"igual à oferta",'
                    '"decidido "&$O%d&"/"&$S%d&"/"&$T%d&" · oferta "&%s&"/"&%s&"/"&%s))')
            % (A, r, bt_(A, 20, "0"), r, bt_(A, 31, "0"), r, bt_(A, 33, "0"),
               r, r, r, bt_(A, 20, "0"), bt_(A, 31, "0"), bt_(A, 33, "0")))
for c, w in ((49, 16), (50, 40)):
    v3.column_dimensions[get_column_letter(c)].width = w
log("V3 · 2 colunas novas: alunos reais 2ª/3ª e divergência ★ x oferta")

# ---- V2: quatro colunas de referência no fim
v2 = wb["Reordenamento 2027 V2"]
for c, h in ((39, "Oferta 2027 ·\n1ª série (turmas)"), (40, "Oferta 2027 ·\n2ª série (turmas)"),
             (41, "Oferta 2027 ·\n3ª série (turmas)"), (42, "★ x Oferta 2027 ·\ndivergência")):
    cabecalho(v2, c, h, ROXO)
for r in range(2, NL + 1):
    A = "$A%d" % r
    v2.cell(r, 39, "=" + bt_(A, 20, "0"))
    v2.cell(r, 40, "=" + bt_(A, 31, "0"))
    v2.cell(r, 41, "=" + bt_(A, 33, "0"))
    v2.cell(r, 42, ('=IF(%s="","",IF(AND($M%d=%s,$Q%d=%s,$R%d=%s),"igual à oferta",'
                    '"decidido "&$M%d&"/"&$Q%d&"/"&$R%d&" · oferta "&%s&"/"&%s&"/"&%s))')
            % (A, r, bt_(A, 20, "0"), r, bt_(A, 31, "0"), r, bt_(A, 33, "0"),
               r, r, r, bt_(A, 20, "0"), bt_(A, 31, "0"), bt_(A, 33, "0")))
for c, w in ((39, 14), (40, 14), (41, 14), (42, 40)):
    v2.column_dimensions[get_column_letter(c)].width = w
log("V2 · 4 colunas novas de referência da oferta 2027")

# ══════════════════════════════════════════════════ 6. backup das ★
print("\n6. Backup das colunas ★ antes desta alteração")
if "★ Backup antes da UETEP" in wb.sheetnames: del wb["★ Backup antes da UETEP"]
bk = wb.create_sheet("★ Backup antes da UETEP")
COLS = [("Reordenamento 2027 V2", 13, "V2 ★1ª SÉRIE"), ("Reordenamento 2027 V2", 24, "V2 ★SALAS"),
        ("Reordenamento 2027 V3", 15, "V3 ★1ª SÉRIE"), ("Reordenamento 2027 V3", 19, "V3 ★2ª SÉRIE"),
        ("Reordenamento 2027 V3", 20, "V3 ★3ª SÉRIE"), ("Reordenamento 2027 V3", 31, "V3 ★SALAS")]
for c, h in enumerate(["INEP", "Escola"] + [x[2] for x in COLS], start=1):
    cabecalho(bk, c, h, "6B7280")
v3vv = wv["Reordenamento 2027 V3"]
for r in range(2, NL + 1):
    i = ip(v3vv.cell(r, 1).value)
    if not i or i == "None": continue
    celula(bk, r, 1, int(numero_(i)))
    celula(bk, r, 2, matrizes[i].escola if i in matrizes else "", centro=False)
    for j, (aba, col, _) in enumerate(COLS):
        celula(bk, r, 3 + j, wv[aba].cell(r, col).value)
for c, w in enumerate([11, 32, 14, 12, 14, 12, 12, 12], start=1):
    bk.column_dimensions[get_column_letter(c)].width = w
bk.row_dimensions[1].height = 34
bk.freeze_panes = "C2"
log("aba '★ Backup antes da UETEP' criada — valores anteriores das 6 colunas ★")

wb.save(SAIDA)
print("\nsalvo:", SAIDA)
pickle.dump(mudancas, open("mudancas.pkl", "wb"))

# ══════════════════════════════════════════════════ 7. Base GDI
# A aba veio de um IMPORTRANGE que o export do Google congelou em
# =IFERROR(__xludf.DUMMYFUNCTION("IMPORTRANGE(...)"), valor).
# Essa fórmula não existe no Sheets: o que aparece é sempre o argumento de
# reserva. Em parte das células a reserva está vazia e o texto só vivia no
# cache do arquivo — que nenhum round-trip preserva. Para não perder nada,
# gravo todas as células como valor e guardo o IMPORTRANGE original à parte.
print("\n7. Base GDI — congelando o IMPORTRANGE como valor, sem perder texto")
gdi_f, gdi_v = wb["Base GDI"], wv["Base GDI"]
importrange = str(gdi_f.cell(1, 1).value or "")
convertidas = 0
for r in range(1, (gdi_v.max_row or 0) + 1):
    for c in range(1, (gdi_v.max_column or 0) + 1):
        f = gdi_f.cell(r, c).value
        if not isinstance(f, str) or not f.startswith("="): continue
        gdi_f.cell(r, c, gdi_v.cell(r, c).value)
        convertidas += 1
log("Base GDI: %d célula(s) de IMPORTRANGE congelado viraram valor" % convertidas)

if "Como restaurar o IMPORTRANGE" in wb.sheetnames: del wb["Como restaurar o IMPORTRANGE"]
nota = wb.create_sheet("Como restaurar o IMPORTRANGE")
nota.column_dimensions["A"].width = 3
nota.column_dimensions["B"].width = 130
texto = [
    ("T", "Base GDI — como voltar a ser um IMPORTRANGE vivo"),
    ("", ""),
    ("L", "Quando uma planilha do Google é baixada em .xlsx, o IMPORTRANGE não vem junto: "
          "vira uma fórmula falsa (__xludf.DUMMYFUNCTION) que só devolve o último valor lido. "
          "Em parte das células nem isso — o texto ficava só no cache do arquivo, que se perde "
          "em qualquer conversão."),
    ("L", "Por isso a aba Base GDI foi gravada como valor: o conteúdo está todo lá, "
          "inclusive as observações longas que se perderiam."),
    ("", ""),
    ("H", "Para voltar a puxar da planilha de origem"),
    ("L", "Apague o conteúdo da aba Base GDI e cole na célula A1:"),
    ("C", importrange[:2000] if importrange.startswith("=") else "(fórmula original não encontrada no arquivo)"),
    ("L", "Na primeira vez o Google pede para autorizar o acesso à planilha de origem."),
    ("", ""),
    ("L", "Enquanto a aba estiver como valor, tudo funciona normalmente — o script lê "
          "valores, não fórmulas. Só não atualiza sozinha quando a origem mudar."),
]
r = 1
for tipo, txt in texto:
    x = nota.cell(r, 2, txt)
    if tipo == "T":
        x.font = Font(name=ARIAL, size=15, bold=True, color="111827"); nota.row_dimensions[r].height = 26
    elif tipo == "H":
        x.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
        x.fill = PatternFill("solid", fgColor="374151")
        x.alignment = Alignment(vertical="center", indent=1); nota.row_dimensions[r].height = 22
    elif tipo == "C":
        x.font = Font(name="Consolas", size=8, color="1F2937")
        x.fill = PatternFill("solid", fgColor="F3F4F6")
        x.alignment = Alignment(wrap_text=True, vertical="top"); nota.row_dimensions[r].height = 60
    else:
        x.font = Font(name=ARIAL, size=10, color="1F2937")
        x.alignment = Alignment(wrap_text=True, vertical="top")
        nota.row_dimensions[r].height = max(15, 14 * (1 + len(txt) // 100))
    r += 1
nota.sheet_view.showGridLines = False
log("aba 'Como restaurar o IMPORTRANGE' criada com a fórmula original")

wb.save(SAIDA)
print("\nsalvo:", SAIDA)
