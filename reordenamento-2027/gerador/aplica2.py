# -*- coding: utf-8 -*-
"""Segunda rodada de ajustes, aplicada sobre o arquivo já entregue."""
import openpyxl, pickle, sys, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gl
from openpyxl.worksheet.datavalidation import DataValidation
sys.path.insert(0, ".")
from motor import *
from agrega import mapa_gdi, processar_turmas, base_ueja, fusao_entre_escolas, panorama
from uetep_motor import ler_oferta, inep_ as ip
from layout_v3 import V3, c, com_quantidade, FUNDAMENTAL, DECISAO_ANEXO, MAX_QTD
from estilo import COR, FUNDO, ARIAL

ENTRADA = "Proposta_Reordenamento_2027_com_UETEP.xlsx"
OFERTA  = "/root/.claude/uploads/41fbd151-f5e2-530d-88f9-9b26c68b08f9/8e7688c9-OFERTA_2027_BASE.xlsx"
SAIDA   = "Proposta_Reordenamento_2027_v4.xlsx"
BORDA = Border(*[Side(style="thin", color="D1D5DB")] * 4)

def cab(ws, col, texto, bloco):
    bg, fg = COR[bloco]
    x = ws.cell(1, col, texto)
    x.fill = PatternFill("solid", fgColor=bg)
    x.font = Font(name=ARIAL, size=8, bold=True, color=fg)
    x.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    x.border = BORDA

def cel(ws, r, col, v, centro=True, wrap=False, fundo=None):
    x = ws.cell(r, col, v)
    x.font = Font(name=ARIAL, size=9)
    x.alignment = Alignment(vertical="top", wrap_text=wrap,
                            horizontal="center" if centro else "general")
    x.border = BORDA
    if fundo: x.fill = PatternFill("solid", fgColor=fundo)

# ═════════════════════════════════════════════ dados
wb = openpyxl.load_workbook(ENTRADA, data_only=False)
wv = openpyxl.load_workbook(ENTRADA, data_only=True)

wo = openpyxl.load_workbook(OFERTA, data_only=True)
oferta, _ = ler_oferta(*[[list(r) for r in wo[a].iter_rows(values_only=True)]
                         for a in ["OFERTA INTEGRAL 2027", "CONTINUIDADE 2027", "SUBSEQUENTE 2027"]])

turmas = [list(r) for r in wv["Turmas"].iter_rows(values_only=True)]
gdi = mapa_gdi([list(r) for r in wv["Base GDI"].iter_rows(values_only=True)])
matrizes, ordem_m, anexos, ordem_a = processar_turmas(turmas, gdi)
ueja_linhas, ueja_resumo = base_ueja(matrizes, ordem_m)
_, fe_resumo = fusao_entre_escolas(matrizes, ordem_m)

from uetep_motor import salas_2027
def salas(i):
    e = matrizes.get(i)
    if not e: return {"salas": 0, "integral": 0, "manha": 0, "tarde": 0, "noite": 0}
    o = oferta.get(i)
    s = salas_2027(e, o)
    if so_eja(e):                       # CEJA: a EJA do prédio matriz é a sala
        s = {"integral": e.ejMatrizI, "manha": e.ejMatrizM,
             "tarde": e.ejMatrizT, "noite": e.ejMatrizN}
        s["salas"] = int(s["integral"] + max(s["manha"], s["tarde"]) + s["noite"])
    return s

mud = []
def log(t): mud.append(t); print("  ·", t)

# ═════════════════════════════════════════════ 1. Validações
print("\n1. Validações — listas com quantidade embutida")
va = wb["Validações"]
cursos27 = [wv["Cursos"].cell(r, 3).value for r in range(2, wv["Cursos"].max_row + 1)
            if str(wv["Cursos"].cell(r, 7).value or "").upper().startswith("OFERTADO")]
novas = [("CURSOS 1ª SÉRIE\ncom quantidade", com_quantidade(cursos27), "em"),
         ("FUNDAMENTAL\ncom quantidade",     com_quantidade(FUNDAMENTAL), "fund"),
         ("DECISÃO SOBRE\nO ANEXO",          DECISAO_ANEXO, "anexo")]
for j, (titulo, itens, bloco) in enumerate(novas):
    col = 11 + j
    cab(va, col, titulo, bloco)
    for r in range(2, va.max_row + 200):
        va.cell(r, col, None)
    for r, it in enumerate(itens, start=2):
        cel(va, r, col, it, centro=False, fundo=FUNDO.get(bloco))
    va.column_dimensions[gl(col)].width = 34
    log("Validações col %s: %s — %d itens" % (gl(col), titulo.replace("\n", " "), len(itens)))
N_CUR, N_FUND, N_ANEXO = [len(x[1]) for x in novas]

# ═════════════════════════════════════════════ 2. Base Tratada
print("\n2. Base Tratada — refeita inteira com as novas regras")
from bases import base_tratada_44, H_BT44
bt = wb["Base Tratada"]
linhas_bt = base_tratada_44(matrizes, ordem_m, anexos, ordem_a, oferta,
                            ueja_resumo, fe_resumo, salas)
for r in range(2, (wv["Base Tratada"].max_row or 2) + 200):
    for col in range(1, 60): bt.cell(r, col, None)
for j, h in enumerate(H_BT44):
    cab(bt, j + 1, h, "eja" if j >= 39 else "id")
LARG = [100, 220, 400, 70, 70, 70, 380, 160, 280, 220, 360, 90, 80, 280, 300, 200,
        100, 110, 130, 110, 400, 300, 110, 120, 400, 320, 80, 130, 180, 110,
        110, 110, 110, 110, 120, 120, 400, 300, 100, 120, 100, 120, 130, 120]
WRAP = {3, 7, 8, 9, 10, 11, 14, 15, 16, 21, 22, 25, 26, 37, 38}
for r, L in enumerate(linhas_bt, start=2):
    for col, v in enumerate(L, start=1):
        cel(bt, r, col, v, centro=(col not in WRAP), wrap=(col in WRAP))
for j, w in enumerate(LARG): bt.column_dimensions[gl(j + 1)].width = w
bt.row_dimensions[1].height = 44
bt.freeze_panes = "C2"
bt.auto_filter.ref = "A1:%s%d" % (gl(44), len(linhas_bt) + 1)
log("Base Tratada: %d linhas x 44 colunas, refeita do zero" % len(linhas_bt))

# ═════════════════════════════════════════════ 2b. Base UETEP
print("\n2b. Base UETEP — refeita com as mesmas regras")
bu = wb["Base UETEP"]
H_BU = ["INEP", "GRE", "Município", "Escola",
        "1ª série 2027 · turmas", "1ª série 2027 · cursos", "1ª série 2027 · alunos previstos",
        "2ª série 2027 · turmas", "2ª série 2027 · alunos", "2ª série 2027 · cursos",
        "3ª série 2027 · turmas", "3ª série 2027 · alunos", "3ª série 2027 · cursos",
        "Subsequente 2027 · turmas", "Subsequente 2027 · alunos", "Subsequente 2027 · cursos",
        "Total de turmas 2027", "Salas necessárias 2027 (oferta real)", "Composição das salas"]
for r in range(2, (wv["Base UETEP"].max_row or 2) + 5):
    for col in range(1, 20): bu.cell(r, col, None)
ordem_of = sorted(oferta, key=lambda i: (oferta[i]["gre"], oferta[i]["municipio"], oferta[i]["escola"]))
for r, i in enumerate(ordem_of, start=2):
    o = oferta[i]; s2 = salas(i) if i in matrizes else None
    vals = [int(i), o["gre"], norm(o["municipio"]), o["escola"],
            int(o["of1_turmas"]), o["of1_detalhe"], int(o["of1_alunos"]),
            int(o["co2_turmas"]), int(o["co2_alunos"]), o["co2_detalhe"],
            int(o["co3_turmas"]), int(o["co3_alunos"]), o["co3_detalhe"],
            int(o["sub_turmas"]), int(o["sub_alunos"]), o["sub_detalhe"],
            int(o["total_turmas"]),
            s2["salas"] if s2 else "",
            ("integral %d + máx(manhã %d, tarde %d) + noite %d"
             % (s2["integral"], s2["manha"], s2["tarde"], s2["noite"])) if s2 else ""]
    for col, v in enumerate(vals, start=1):
        cel(bu, r, col, v, centro=(col not in (3, 4, 6, 10, 13, 16, 19)),
            wrap=(col in (6, 10, 13, 16, 19)))
bu.auto_filter.ref = "A1:S%d" % (len(ordem_of) + 1)
log("Base UETEP: %d escolas, refeita" % len(ordem_of))

# ═════════════════════════════════════════════ 3. Base UEJA
print("\n3. Base UEJA")
uj = wb["Base UEJA"]
antigas = {}
for r in range(2, wv["Base UEJA"].max_row + 1):
    k = (ip(wv["Base UEJA"].cell(r, 1).value),
         str(wv["Base UEJA"].cell(r, 5).value or "").upper().strip(),
         str(wv["Base UEJA"].cell(r, 6).value or "").upper().strip())
    antigas[k] = wv["Base UEJA"].cell(r, 11).value
for r in range(2, uj.max_row + 200):
    for col in range(1, 12): uj.cell(r, col, None)
for r, L in enumerate(ueja_linhas, start=2):
    k = (str(L[0]), str(L[4]).upper().strip(), str(L[5]).upper().strip())
    L = list(L)[:10] + [antigas.get(k, "")]
    for col, v in enumerate(L, start=1):
        cel(uj, r, col, v, centro=(col not in (3, 4, 5, 6, 9, 11)), wrap=(col in (5, 6, 9, 11)))
if uj.max_row > len(ueja_linhas) + 1:
    uj.delete_rows(len(ueja_linhas) + 2, uj.max_row - len(ueja_linhas) - 1)
uj.auto_filter.ref = "A1:K%d" % (len(ueja_linhas) + 1)
log("Base UEJA: %d linhas, já sem as turmas zeradas" % len(ueja_linhas))

# ═════════════════════════════════════════════ 4. Reordenamento 2027 V3
print("\n4. Reordenamento 2027 V3 — novo layout")
antigo, antigo_v = wb["Reordenamento 2027 V3"], wv["Reordenamento 2027 V3"]

# guarda tudo o que foi digitado, por INEP e por rótulo de coluna
def rotulo(s): return re.sub(r"\s+", " ", str(s or "")).strip().upper()
heads_old = {rotulo(antigo_v.cell(1, col).value): col
             for col in range(1, antigo_v.max_column + 1)}
guardado = {}
for r in range(2, antigo_v.max_row + 1):
    i = ip(antigo_v.cell(r, 1).value)
    if not i or i == "None": continue
    for h, col in heads_old.items():
        f = antigo.cell(r, col).value
        if isinstance(f, str) and f.startswith("="): continue
        v = antigo_v.cell(r, col).value
        if v in (None, "") or str(v).strip() == "": continue
        guardado.setdefault(i, {})[h] = v

# ─────────────────────────────────────────────────────────────────────
# Semente x decisão
#
# O script precisa saber se um número numa coluna ★ foi ele mesmo que
# plantou ou se alguém digitou. Guardar isso na memória de quem roda não
# funciona — então fica gravado na aba "★ Sementes (controle)": a cada
# rodada o script anota o que semeou. Célula que ainda tem exatamente a
# última semente pode ser refeita; qualquer outra coisa é decisão humana
# e não se toca.
# ─────────────────────────────────────────────────────────────────────
ABA_SEM = "★ Sementes (controle)"
SEM_ANTERIOR = {}
if ABA_SEM in wv.sheetnames:
    sh = wv[ABA_SEM]
    cols = {str(sh.cell(1, x).value or "").strip(): x for x in range(2, sh.max_column + 1)}
    for r in range(2, sh.max_row + 1):
        i_ = ip(sh.cell(r, 1).value)
        if not i_ or i_ == "None": continue
        SEM_ANTERIOR[i_] = {h: sh.cell(r, x).value for h, x in cols.items()}
else:
    # primeira vez: a entrega anterior semeou tudo, menos o que ela preservou
    import os.path
    preservadas = pickle.load(open("preservadas.pkl", "rb")) if os.path.exists("preservadas.pkl") else []
    decisoes = {(x[0], x[2]) for x in preservadas}
    for r in range(2, antigo_v.max_row + 1):
        i_ = ip(antigo_v.cell(r, 1).value)
        if not i_ or i_ == "None": continue
        for h, col in heads_old.items():
            v_ = antigo_v.cell(r, col).value
            if ("Reordenamento 2027 V3", i_) in decisoes and "1ª SÉRIE" in h:
                v_ = None                      # foi decisão, não semente
            SEM_ANTERIOR.setdefault(i_, {})[h] = v_
    print("   (primeira rodada: semente reconstruída da entrega anterior, "
          "%d decisão(ões) marcada(s))" % len(decisoes))

SEMEADO = {}          # o que ESTA rodada semear, para gravar no controle

DE_PARA = {
    "★ FUNDAMENTAL 2027 DECIDA AQUI":            "★ FUNDAMENTAL",
    "★ 1ª SÉRIE 2027 DECIDA AQUI":               "★ 1ª SÉRIE",
    "★ CURSOS 1ª SÉRIE 2027 DECIDA AQUI":        "★ CURSOS 1ª SÉRIE",
    "★ ALTERAÇÃO DE CURSOS EMI":                 "★ ALTERAÇÃO DE",
    "★ 2ª SÉRIE 2027 DECIDA AQUI":               "★ 2ª SÉRIE",
    "★ 3ª SÉRIE 2027 DECIDA AQUI":               "★ 3ª SÉRIE",
    "★ PARCIAIS 2026 JUSTIFICATIVA":             "★ Parciais 2026",
    "★ VALIDAÇÃO DA FUSÃO":                      "★ Validação",
    "★ SALAS NECESSÁRIAS 2027 — DECIDA AQUI":    "★ SALAS",
    "★ EJA TURMAS 2027":                         "★ EJA TURMAS",
    "★ REORDENAMENTO 2027":                      "★ Reordenamento",
    "★ JUSTIFICATIVA":                           "★ Justificativa",
    "★ OBSERVAÇÃO":                              "★ Observação",
    "✓ PRONTO":                                  "✓ Pronto",
}
ANEXO_ANTIGAS = ["★ MANTER O ANEXO?", "★ ENCERRAR O ANEXO?", "★ REMANEJAR A OFERTA DO ANEXO?"]

del wb["Reordenamento 2027 V3"]
v3 = wb.create_sheet("Reordenamento 2027 V3")
NL = len(ordem_m) + 1
BT = "'Base Tratada'"; PM = "'Panorama Municipal'"; GD = "'Base GDI'"
def ultima_com_conteudo(aba):
    """Igual ao getLastRow() do Sheets: última linha com algo em QUALQUER coluna."""
    ws = wv[aba]
    ultima = 2
    for r, linha in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v not in (None, "") and str(v).strip() for v in linha):
            ultima = r
    return max(2, ultima)

NBT = len(linhas_bt) + 1
NPM, NGD = ultima_com_conteudo("Panorama Municipal"), ultima_com_conteudo("Base GDI")

def bt_(a, col, err='""'):
    return "IFERROR(VLOOKUP(%s,%s!$A$2:$AR$%d,%d,FALSE),%s)" % (a, BT, NBT, col, err)
def pm_(a, col, err='""'):
    return "IFERROR(VLOOKUP(%s,%s!$A$2:$N$%d,%d,FALSE),%s)" % (a, PM, NPM, col, err)
def gd_(a, col, err='""'):
    return "IFERROR(VLOOKUP(%s,%s!$A$2:$AL$%d,%d,FALSE),%s)" % (a, GD, NGD, col, err)
def soma_parenteses(cel_ref):
    """Lê 'Logística (2), Agropecuária (1)' e devolve 3. Nunca fica numa célula ★."""
    return ('IFERROR(SUM(ARRAYFORMULA(IFERROR(VALUE(REGEXEXTRACT('
            'SPLIT(%s,","),"\\((\\d+)\\)")),0))),0)' % cel_ref)

for col, (nome, bloco, larg) in enumerate(V3, start=1):
    cab(v3, col, nome, bloco)
    v3.column_dimensions[gl(col)].width = larg
v3.row_dimensions[1].height = 64
v3.freeze_panes = "E2"

ineps = [str(L[0]) for L in linhas_bt]
L = gl                       # atalho: número da coluna → letra
def ref(chave, r): return "$%s%d" % (gl(c(chave)), r)

for k, i in enumerate(ineps):
    r = k + 2
    A, C = "$A%d" % r, "$C%d" % r
    e = matrizes.get(i); o = oferta.get(i); s_ = salas(i)
    g = guardado.get(i, {})

    def put(chave, v): v3.cell(r, c(chave), v)

    put("inep", int(numero_(i)))
    put("gre", "=" + gd_(A, 3))
    put("municipio", norm(e.municipio) if e else "")
    put("escola", "=" + gd_(A, 5))
    put("turmas_hoje", "=" + gd_(A, 14))
    put("turmas_2027", "=" + gd_(A, 15))
    put("oferta_fund", "=" + bt_(A, 9, '"—"'))
    put("fund_total", "=" + soma_parenteses(ref("fund_turmas", r)))
    put("ano9_escola", '=IF(%s="","",%s&" matrícula(s) · "&%s&" turma(s)")'
        % (A, bt_(A, 18, "0"), bt_(A, 17, "0")))
    put("ano9_municipio", '=IF(%s="","",%s&" alunos (estadual "&%s&" + outras redes "&%s&")")'
        % (C, pm_(C, 2, "0"), pm_(C, 4, "0"), pm_(C, 6, "0")))
    put("demanda", '=IF(%s="","",%s&" alunos → "&%s&" turma(s) necessária(s)")'
        % (C, pm_(C, 7, "0"), pm_(C, 8, "0")))
    saldo = pm_(C, 13, '""')
    put("faltam", '=IF(%s="","",IF(%s="","—",IF(%s>0,"FALTAM DISTRIBUIR "&%s&" TURMA(S)",'
        'IF(%s=0,"DEMANDA MUNICIPAL ATENDIDA","EXCEDE A DEMANDA EM "&ABS(%s)&" TURMA(S)"))))'
        % (C, saldo, saldo, saldo, saldo, saldo))
    put("ideb", "=IFERROR(VLOOKUP($A%d,\'IDEB - ESCOLAS\'!C:F,4,0),\"\")" % r)
    put("s1_2026", int(e.s1) if e else 0)
    put("cursos_2026", "=" + bt_(A, 3, '"—"'))
    put("cursos_total", "=" + soma_parenteses(ref("cursos_turmas", r)))
    put("of_1a", "=" + bt_(A, 20, "0"))
    put("of_cursos", "=" + bt_(A, 21, '"—"'))
    put("of_23sub", '=IF(%s="","",%s&" de 2ª | "&%s&" de 3ª | "&%s&" subseq.")'
        % (A, bt_(A, 31, "0"), bt_(A, 33, "0"), bt_(A, 35, "0")))
    put("parciais", "=" + bt_(A, 10, '"—"'))
    put("fusao_propria", "=" + bt_(A, 11, '"—"'))
    put("fusao_entre", "=" + bt_(A, 25, '"—"'))
    put("salas_existentes", "=" + gd_(A, 6))

    # ---- salas: a base do motor, movida por tudo que for decidido nas ★
    turmas1a = "(IF(%s>0,%s,%s))" % (ref("cursos_total", r), ref("cursos_total", r), ref("s1_2027", r))
    put("salas_calc", '=IF(%s="","",MAX(0,%s+(%s-%s)+(%s-%s)+(%s-%s)+(%s-%s)+IF(%s="SIM",%s-%s,0)))'
        % (A, bt_(A, 28, "0"),
           turmas1a, bt_(A, 20, "0"),
           ref("s2_2027", r), bt_(A, 31, "0"),
           ref("s3_2027", r), bt_(A, 33, "0"),
           ref("fund_total", r), bt_(A, 40, "0"),
           bt_(A, 41, '"NÃO"'), ref("eja_turmas", r), bt_(A, 42, "0")))
    put("salas_situacao",
        '=IF(%s="","",IF(%s>%s,(%s-%s)&" SALA(S) OCIOSA(S)",IF(%s<%s,"CONSTRUIR "&(%s-%s)&" SALA(S)",'
        '"QUANTIDADE ADEQUADA")))'
        % (ref("salas_decide", r), ref("salas_existentes", r), ref("salas_decide", r),
           ref("salas_existentes", r), ref("salas_decide", r),
           ref("salas_existentes", r), ref("salas_decide", r),
           ref("salas_decide", r), ref("salas_existentes", r)))

    put("eja_oferta", "=" + bt_(A, 7, '"—"'))
    put("eja_matriculas", "=" + bt_(A, 16, '"—"'))
    put("eja_fora", "=" + bt_(A, 22, '"—"'))
    put("anexos_lista", "=" + bt_(A, 26, '"—"'))
    put("anexos_qtd", "=" + bt_(A, 27, "0"))
    put("resumo", '=IF(%s="","",%s&" de 1ª | "&%s&" de 2ª | "&%s&" de 3ª | Fund "&%s&" | EJA "&%s'
        '&" | Necessidade: "&%s&" salas | "&%s)'
        % (A, turmas1a, ref("s2_2027", r), ref("s3_2027", r), ref("fund_total", r),
           ref("eja_turmas", r), ref("salas_decide", r), ref("salas_situacao", r)))
    put("escola_proxima", "=" + gd_(A, 18, '"—"'))
    put("of_alunos23", '=IF(%s="","",%s+%s&" aluno(s)")' % (A, bt_(A, 32, "0"), bt_(A, 34, "0")))
    put("divergencia", '=IF(%s="","",IF(AND(%s=%s,%s=%s,%s=%s),"igual à oferta",'
        '"decidido "&%s&"/"&%s&"/"&%s&" · oferta "&%s&"/"&%s&"/"&%s))'
        % (A, turmas1a, bt_(A, 20, "0"), ref("s2_2027", r), bt_(A, 31, "0"),
           ref("s3_2027", r), bt_(A, 33, "0"),
           turmas1a, ref("s2_2027", r), ref("s3_2027", r),
           bt_(A, 20, "0"), bt_(A, 31, "0"), bt_(A, 33, "0")))

    # ---- colunas ★: valor digitado antes tem prioridade; senão, semente
    def num(x):
        try: return round(float(x), 4)
        except (TypeError, ValueError): return None

    def estrela(chave, antigo_rot, semente=None):
        """Decisão digitada sempre ganha. Semente antiga é refeita pela nova."""
        v = g.get(antigo_rot)
        vazio = v in (None, "") or str(v).strip() == ""
        if not vazio and semente is not None:
            anterior = SEM_ANTERIOR.get(i, {}).get(antigo_rot)
            if num(v) is not None and num(v) == num(anterior):
                v, vazio = None, True          # continuava sendo semente
        if vazio: v = semente
        if semente is not None:
            SEMEADO.setdefault(i, {})[antigo_rot] = semente
        if v not in (None, ""): v3.cell(r, c(chave), v)

    estrela("fund_etapas",  "★ FUNDAMENTAL 2027 DECIDA AQUI")
    estrela("fund_turmas",  "★ TURMAS FUNDAMENTAL 2027 — DECIDA AQUI")
    estrela("s1_2027",      "★ 1ª SÉRIE 2027 DECIDA AQUI", int(o["of1_turmas"]) if o else 0)
    estrela("cursos_2027",  "★ CURSOS 1ª SÉRIE 2027 DECIDA AQUI")
    estrela("cursos_turmas","★ TURMAS 1ª SÉRIE 2027 DECIDA AQUI")
    estrela("alteracao_emi","★ ALTERAÇÃO DE CURSOS EMI")
    estrela("s2_2027",      "★ 2ª SÉRIE 2027 DECIDA AQUI", int(o["co2_turmas"]) if o else 0)
    estrela("s3_2027",      "★ 3ª SÉRIE 2027 DECIDA AQUI", int(o["co3_turmas"]) if o else 0)
    estrela("parciais_just","★ PARCIAIS 2026 JUSTIFICATIVA")
    estrela("fusao_valid",  "★ VALIDAÇÃO DA FUSÃO")
    estrela("salas_decide", "★ SALAS NECESSÁRIAS 2027 — DECIDA AQUI", s_["salas"])
    estrela("eja_turmas",   "★ EJA TURMAS 2027", int(e.ejaMatrizTurmas) if e else 0)
    estrela("reord",        "★ REORDENAMENTO 2027")
    estrela("justificativa","★ JUSTIFICATIVA")
    estrela("observacao",   "★ OBSERVAÇÃO")
    estrela("pronto",       "✓ PRONTO")

    # as três perguntas de anexo viram uma só
    juntas = [g.get(x) for x in ANEXO_ANTIGAS]
    juntas = [str(x).strip() for x in juntas if x not in (None, "") and str(x).strip()]
    if juntas: v3.cell(r, c("anexos_decisao"), " · ".join(juntas))

# formatação do corpo
for r in range(2, NL + 1):
    for col, (nome, bloco, _) in enumerate(V3, start=1):
        x = v3.cell(r, col)
        x.font = Font(name=ARIAL, size=9, bold=(bloco == "estrela"),
                      color="92400E" if bloco == "estrela" else "000000")
        x.alignment = Alignment(vertical="top", wrap_text=(V3[col-1][2] >= 20),
                                horizontal="general" if V3[col-1][2] >= 20 else "center")
        x.border = BORDA
        if bloco in FUNDO: x.fill = PatternFill("solid", fgColor=FUNDO[bloco])
v3.auto_filter.ref = "A1:%s%d" % (gl(len(V3)), NL)
log("V3 refeita: %d colunas · %d linhas" % (len(V3), NL - 1))

# ---- menus
def menu(chave, aba, col, n, titulo_="", msg=""):
    dv = DataValidation(type="list", allow_blank=True, showDropDown=False,
                        formula1="'%s'!$%s$2:$%s$%d" % (aba, gl(col), gl(col), n + 1))
    dv.promptTitle, dv.prompt, dv.showInputMessage = titulo_, msg, bool(msg)
    v3.add_data_validation(dv)
    letra = gl(c(chave))
    dv.add("%s2:%s%d" % (letra, letra, NL))

menu("fund_etapas",   "Validações", 1,  11, "Fundamental 2027", "Quais etapas a escola mantém.")
menu("fund_turmas",   "Validações", 12, N_FUND, "Turmas de Fundamental",
     "Escolha a etapa já com a quantidade. Clicar de novo acrescenta outra.")
menu("cursos_2027",   "Validações", 11, N_CUR, "Cursos da 1ª série")
menu("cursos_turmas", "Validações", 11, N_CUR, "Turmas por curso",
     "Escolha o curso já com a quantidade. Clicar de novo acrescenta outro.")
menu("fusao_valid",   "Validações", 8,  7, "Validação da fusão")
menu("anexos_decisao","Validações", 13, N_ANEXO, "Decisão sobre o anexo",
     "Uma escolha só: manter, encerrar ou remanejar.")
menu("reord",         "Validações", 9,  7, "Reordenamento 2027")
log("menus ligados — curso e etapa já vêm com a quantidade, nada é digitado")

# ═════════════════════════════════════════════ 5. aba de controle das sementes
print("\n5. Registro das sementes")
if ABA_SEM in wb.sheetnames: del wb[ABA_SEM]
ctl = wb.create_sheet(ABA_SEM)
rots = ["★ 1ª SÉRIE 2027 DECIDA AQUI", "★ 2ª SÉRIE 2027 DECIDA AQUI",
        "★ 3ª SÉRIE 2027 DECIDA AQUI", "★ SALAS NECESSÁRIAS 2027 — DECIDA AQUI",
        "★ EJA TURMAS 2027"]
cab(ctl, 1, "INEP", "id")
for j, h in enumerate(rots): cab(ctl, 2 + j, h, "estrela")
for k, i in enumerate(ineps):
    cel(ctl, k + 2, 1, int(numero_(i)))
    for j, h in enumerate(rots):
        cel(ctl, k + 2, 2 + j, SEMEADO.get(i, {}).get(h))
ctl.column_dimensions["A"].width = 11
for j in range(len(rots)): ctl.column_dimensions[gl(2 + j)].width = 18
ctl.row_dimensions[1].height = 40
ctl.freeze_panes = "B2"
ctl.sheet_state = "hidden"
log("aba '%s' gravada — o script usa isso para não confundir semente com decisão" % ABA_SEM)

wb.save(SAIDA)
print("\nsalvo:", SAIDA)
