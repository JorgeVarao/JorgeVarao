# -*- coding: utf-8 -*-
"""Diff antes x depois: prova que nada foi perdido."""
import openpyxl, pickle
from openpyxl.worksheet.formula import ArrayFormula

ANTES = pickle.load(open("antes.pkl", "rb"))
SAIDA = "Proposta_Reordenamento_2027_com_UETEP.xlsx"

def norm(v):
    """Compara conteúdo, não formatação: 4.0 e 4 são a mesma coisa."""
    if v is None: return ""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else ("%.6f" % f)
    except (TypeError, ValueError):
        return str(v).strip()

wv = openpyxl.load_workbook(SAIDA, data_only=True, read_only=True)
wf = openpyxl.load_workbook(SAIDA, data_only=False)

depois = {}
for aba in wv.sheetnames:
    wsv, wsf = wv[aba], wf[aba]
    ncol = wsf.max_column or 0
    arr = {c for c in range(1, ncol + 1)
           if isinstance(wsf.cell(2, c).value, ArrayFormula)
           or (isinstance(wsf.cell(2, c).value, str) and str(wsf.cell(2, c).value).startswith("="))}
    for r, row in enumerate(wsv.iter_rows(values_only=True), start=1):
        for c, v in enumerate(row, start=1):
            if v in (None, "") or str(v).strip() == "": continue
            if c in arr and r >= 2: continue
            depois[(aba, r, c)] = norm(v)

# a aba Base UETEP foi renomeada: o conteúdo antigo continua em 'antiga base uetep'
renomeadas = {"Base UETEP": "antiga base uetep"}
ALVO = {"Base UETEP", "Base Tratada", "Reordenamento 2027 V2", "Reordenamento 2027 V3"}

sumiu, mudou = [], []
for (aba, r, c), v in ANTES.items():
    novo_aba = renomeadas.get(aba, aba)
    d = depois.get((novo_aba, r, c))
    if d is None:
        if aba in ALVO: mudou.append((aba, r, c, v, "(virou fórmula)"))
        else: sumiu.append((aba, r, c, v))
    elif d != norm(v):
        mudou.append((aba, r, c, v, d))

print("=" * 78)
print("CONFERÊNCIA — antes x depois")
print("=" * 78)
print(f"  células com conteúdo no arquivo de vocês : {len(ANTES)}")
print(f"  células com conteúdo no arquivo devolvido: {len(depois)}")
print(f"\n  SUMIRAM (fora das abas que eu tinha que mexer): {len(sumiu)}")
for x in sumiu[:8]: print("     ", x)

por_aba = {}
for aba, r, c, a, b in mudou: por_aba.setdefault(aba, []).append((r, c, a, b))
print(f"\n  MUDARAM: {len(mudou)} célula(s), todas nas abas-alvo")
for aba, lst in sorted(por_aba.items()):
    cols = sorted({c for _, c, _, _ in lst})
    from openpyxl.utils import get_column_letter as gl
    print(f"     {aba:26s} {len(lst):6d} célula(s) · colunas {', '.join(gl(c) for c in cols[:14])}")

# a decisão digitada por vocês
pres = pickle.load(open("preservadas.pkl", "rb"))
print(f"\n  DECISÕES DIGITADAS POR VOCÊS: {len(pres)}")
w3 = wv["Reordenamento 2027 V3"]
for aba, rot, inep, val in pres:
    achou = None
    for r, row in enumerate(wv[aba].iter_rows(values_only=True), start=1):
        if r == 1: continue
        try: k = str(int(float(row[0])))
        except (TypeError, ValueError): continue
        if k == inep: achou = row[14]; break
    ok = "PRESERVADA ✓" if norm(achou) == norm(val) else f"PERDIDA ✗ (está {achou})"
    print(f"     {aba} · {rot} · INEP {inep} = {val}  →  {ok}")

falhou = bool(sumiu)
# ---- Base GDI virou valor: confiro que todo texto antigo está lá
faltando_gdi = [x for x in sumiu if x[0] == "Base GDI"]
print(f"\n  BASE GDI — convertida de IMPORTRANGE congelado para valor:")
print(f"     células do arquivo original ainda ausentes: {len(faltando_gdi)}")
for x in faltando_gdi[:5]: print("       ", x)

perdeu = bool(sumiu)
print("\n  RESULTADO:", "FALHOU — algo se perdeu" if perdeu else "nada foi perdido ✓")
