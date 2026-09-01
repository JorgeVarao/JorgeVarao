# -*- coding: utf-8 -*-
"""Estilos compartilhados pelas abas geradas."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ARIAL   = "Arial"
BORDA   = Border(*[Side(style="thin", color="D1D5DB")] * 4)

# paleta dos blocos (mesma lógica de cores do script atual)
COR = {
    "id":      ("374151", "FFFFFF"),   # cinza escuro  — identificação
    "hoje":    ("1E40AF", "FFFFFF"),   # azul          — retrato de 2026
    "fund":    ("0891B2", "FFFFFF"),   # ciano         — fundamental / 9º ano
    "em":      ("059669", "FFFFFF"),   # verde         — ensino médio / séries
    "uetep":   ("0F766E", "FFFFFF"),   # verde-azulado — UETEP
    "fusao":   ("7C3AED", "FFFFFF"),   # roxo          — fusões
    "sala":    ("D97706", "FFFFFF"),   # âmbar         — salas
    "eja":     ("B45309", "FFFFFF"),   # marrom        — EJA
    "anexo":   ("9333EA", "FFFFFF"),   # violeta       — anexos
    "saida":   ("6B7280", "FFFFFF"),   # cinza         — saída / resumo
    "decisao": ("DC2626", "FFFFFF"),   # vermelho      — registro da decisão
    "estrela": ("FBBF24", "000000"),   # amarelo       — ★ preencher à mão
}
FUNDO = {
    "fund":  "ECFEFF", "em": "ECFDF5", "eja": "FEF3C7", "sala": "FFF7ED",
    "uetep": "F0FDFA", "fusao": "F5F3FF", "anexo": "FAF5FF", "estrela": "FEF9C3",
    "saida": "F9FAFB",
}

def cabecalho(ws, headers, blocos, altura=58):
    """headers = lista de títulos; blocos = lista do mesmo tamanho com a chave de cor."""
    for j, (h, b) in enumerate(zip(headers, blocos), start=1):
        c = ws.cell(1, j, h)
        bg, fg = COR[b]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name=ARIAL, size=8, bold=True, color=fg)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDA
    ws.row_dimensions[1].height = altura
    ws.freeze_panes = "A2"

def corpo(ws, nlin, ncol, larguras=None, wrap=None, centro=None, fundos=None):
    if nlin > 0:
        for r in range(2, nlin + 2):
            for j in range(1, ncol + 1):
                c = ws.cell(r, j)
                c.font = Font(name=ARIAL, size=9)
                c.alignment = Alignment(vertical="top",
                                        wrap_text=bool(wrap and j in wrap),
                                        horizontal="center" if (centro and j in centro) else "general")
                c.border = BORDA
        if fundos:
            for j, chave in fundos.items():
                fill = PatternFill("solid", fgColor=FUNDO[chave])
                for r in range(2, nlin + 2):
                    ws.cell(r, j).fill = fill
    if larguras:
        for j, w in larguras.items():
            ws.column_dimensions[get_column_letter(j)].width = w
    if nlin > 0:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ncol), nlin + 1)

def dropdown(ws, col, nlin, itens=None, formula=None, titulo="", msg=""):
    if nlin < 1: return None
    f = formula if formula else '"%s"' % ",".join(itens)
    dv = DataValidation(type="list", formula1=f, allow_blank=True, showDropDown=False)
    dv.promptTitle = titulo or ""
    dv.prompt = msg or ""
    dv.showInputMessage = bool(msg)
    ws.add_data_validation(dv)
    L = get_column_letter(col)
    dv.add("%s2:%s%d" % (L, L, nlin + 1))
    return dv

def titulo_aba(ws, texto, ncol, sub=""):
    """Faixa de título acima do cabeçalho (usada nas abas auxiliares novas)."""
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1, 1, texto + (("  —  " + sub) if sub else ""))
    c.fill = PatternFill("solid", fgColor="111827")
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 24
