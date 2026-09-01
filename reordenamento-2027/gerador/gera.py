# -*- coding: utf-8 -*-
"""
GERADOR DO ARQUIVO ENTREGUE
Reordenamento 2027 — SEDUC-PI / SUPEX / UGERF / GDI
"""
import math, os, pickle
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from motor import *
from agrega import *
from bases import *
from estilo import *
import reord

LIMITE = int(os.environ.get("LIMITE", "0"))   # >0 gera uma amostra para teste
SAIDA = os.environ.get("SAIDA", "Reordenamento_2027_v1_v2_v3.xlsx")

# ==================================================================== dados
d = carregar()
gdi = mapa_gdi(d["SEDUC:Base GDI"])
matrizes, ordem_m, anexos, ordem_a = processar_turmas(d["SEDUC:Turmas"], gdi)
if LIMITE:
    ordem_m = ordem_m[:LIMITE]
    vivos = set(ordem_m)
    matrizes = {k: v for k, v in matrizes.items() if k in vivos}
    ordem_a = [c for c in ordem_a if anexos[c].inep in vivos]
    anexos = {k: v for k, v in anexos.items() if k in set(ordem_a)}
uetep_linhas, uetep_idx = base_uetep(d["REORD:Matriculas por etapa"], gdi)
if LIMITE:
    uetep_linhas = [L for L in uetep_linhas if L[0] in vivos]
    uetep_idx = {k: v for k, v in uetep_idx.items() if k in vivos}
ueja_linhas, ueja_resumo = base_ueja(matrizes, ordem_m)
fe_linhas, fe_resumo = fusao_entre_escolas(matrizes, ordem_m)
pan_linhas, pan_idx = panorama(d["SEDUC:Panorama Municipal"], matrizes, ordem_m)
if LIMITE:
    muns = set(norm(matrizes[i].municipio) for i in ordem_m)
    pan_linhas = [L for L in pan_linhas if norm(L[0]) in muns]

# municípios normalizados em toda a planilha
for i in ordem_m:
    matrizes[i].municipio = norm(matrizes[i].municipio)
for ch in ordem_a:
    anexos[ch].municipio = norm(anexos[ch].municipio)
for L in pan_linhas: L[0] = norm(L[0])
for L in fe_linhas:  L[1] = norm(L[1])
for L in uetep_linhas: L[2] = norm(L[2])
for L in ueja_linhas:  L[2] = norm(L[2])

bt_linhas = base_tratada(matrizes, ordem_m, anexos, ordem_a,
                         uetep_idx, ueja_resumo, fe_resumo)
ba_linhas = base_anexos(anexos, ordem_a)

# ordem de INEP da aba Reordenamento (mesma da planilha original)
ineps_orig = [inep_(r[0]) for r in d["SEDUC:Reordenamento 2027"][1:578] if inep_(r[0])]
ineps = ineps_orig if len(ineps_orig) == len(ordem_m) else list(ordem_m)
if LIMITE:
    ineps = [i for i in ineps_orig if i in set(ordem_m)] or list(ordem_m)

N_BT, N_PM, N_GD = len(bt_linhas), len(pan_linhas), len(d["SEDUC:Base GDI"]) - 1
N_FE = len(fe_linhas)
municipios = [L[0] for L in pan_linhas]
N_ID = len(municipios)
N_FUS = len(d["SEDUC:Fusão Turmas"]) - 1

wb = openpyxl.Workbook()
wb.remove(wb.active)

def escreve(ws, headers, blocos, linhas, larguras=None, wrap=None, centro=None,
            fundos=None, altura=58):
    cabecalho(ws, headers, blocos, altura)
    for r, L in enumerate(linhas, start=2):
        for c, v in enumerate(L, start=1):
            ws.cell(r, c, v)
    corpo(ws, len(linhas), len(headers), larguras, wrap, centro, fundos)

# ==================================================================== LEIA-ME
ws = wb.create_sheet("LEIA-ME")

# ============================================================ REORDENAMENTO
def monta_reord(nome, headers, blocos, formulas, valores, larguras, wrap, centro,
                estrelas):
    ws = wb.create_sheet(nome)
    cabecalho(ws, headers, blocos, 62)
    n = len(ineps)
    for k in range(n):
        r = k + 2
        for c, v in (valores[k] or {}).items():
            ws.cell(r, c, v)
        for c, v in (formulas[k] if formulas else {}).items():
            ws.cell(r, c, v)
    fundos = {}
    for j, b in enumerate(blocos, start=1):
        if b in FUNDO: fundos[j] = b
    corpo(ws, n, len(headers), larguras, wrap, centro, fundos)
    ws.freeze_panes = "E2"
    for j in estrelas:
        for r in range(2, n + 2):
            ws.cell(r, j).font = Font(name=ARIAL, size=9, bold=True, color="92400E")
    return ws

# ---------------------------------------------------------------- V1
R1 = d["SEDUC:Reordenamento 2027"]
val_v1 = []
for k in range(len(ineps)):
    src = R1[k + 1]
    f = {}
    for c in range(1, len(reord.H_V1) + 1):
        v = src[c - 1] if c - 1 < len(src) else None
        if v in (None, ""): continue
        if c == 3: v = norm(v)
        if c == 1: v = int(numero_(v))
        f[c] = v
    val_v1.append(f)

LARG_V1 = {1: 11, 2: 13, 3: 20, 4: 30, 5: 34, 6: 34, 7: 26, 8: 30, 9: 26, 10: 22,
           11: 12, 12: 9, 13: 13, 14: 48, 15: 26, 16: 22, 17: 8, 18: 8, 19: 26,
           20: 22, 21: 38, 22: 26, 23: 10, 24: 15, 25: 20, 26: 44, 27: 20, 28: 13,
           29: 26, 30: 32, 31: 32, 32: 10, 33: 34, 34: 26, 35: 18, 36: 26, 37: 26, 38: 9}
WRAP_V1 = {5, 6, 7, 8, 9, 10, 14, 15, 16, 19, 20, 21, 22, 26, 27, 29, 30, 31, 33, 34, 35, 36, 37}
CENT_V1 = {1, 2, 12, 13, 17, 18, 23, 24, 25, 28, 38}
EST_V1  = [10, 13, 15, 16, 20, 22, 24, 28, 29, 35, 36, 37, 38]

ws1 = monta_reord("Reordenamento 2027", reord.H_V1, reord.B_V1, None, val_v1,
                  LARG_V1, WRAP_V1, CENT_V1, EST_V1)
dropdown(ws1, 10, len(ineps), formula="'Validações'!$A$2:$A$12")
dropdown(ws1, 25, len(ineps), itens=["CONSTRUIR", "OK"])
dropdown(ws1, 35, len(ineps), itens=["ABSORVER", "Opção 2"])

# ---------------------------------------------------------------- V2
val_v2, i_bt = [], {str(L[0]): L for L in bt_linhas}
for k, i in enumerate(ineps):
    L = i_bt.get(i)
    f = {1: int(numero_(i)), 3: matrizes[i].municipio if i in matrizes else ""}
    if L:
        f[12] = L[3]                    # 1ª Série 2026
        f[13] = L[3]                    # ★ 1ª SÉRIE 2027 — parte da projeção
        f[17] = L[4]                    # 2ª Série 2027
        f[18] = L[5]                    # 3ª Série 2027
        f[24] = L[11]                   # ★ SALAS NECESSÁRIAS 2027 (motor)
    val_v2.append(f)

frm_v2 = reord.linhas_v2(ineps, N_BT, N_PM, N_GD, N_ID, n_fus=N_FUS)

ws2 = monta_reord("Reordenamento 2027 V2", reord.H_V2, reord.B_V2, frm_v2, val_v2,
                  LARG_V1, WRAP_V1, CENT_V1, EST_V1)
dropdown(ws2, 10, len(ineps), formula="'Validações'!$A$2:$A$12",
         titulo="Fundamental 2027", msg="Etapas do Ensino Fundamental que a escola manterá em 2027.")
dropdown(ws2, 22, len(ineps), formula="'Validações'!$H$2:$H$8",
         titulo="Validação da fusão", msg="Confirme, recuse ou adie a fusão sugerida.")
dropdown(ws2, 35, len(ineps), formula="'Validações'!$I$2:$I$8",
         titulo="Reordenamento 2027", msg="Decisão final sobre a unidade.")

# ---------------------------------------------------------------- V3
val_v3 = []
for k, i in enumerate(ineps):
    L = i_bt.get(i)
    f = {1: int(numero_(i)), 3: matrizes[i].municipio if i in matrizes else ""}
    if L:
        f[14] = L[3]        # 1ª Série 2026
        f[15] = L[3]        # ★ 1ª SÉRIE 2027
        f[19] = L[4]        # ★ 2ª SÉRIE 2027 (editável)
        f[20] = L[5]        # ★ 3ª SÉRIE 2027 (editável)
        f[31] = L[27]       # ★ SALAS NECESSÁRIAS 2027 (já com UETEP)
    val_v3.append(f)

frm_v3 = reord.linhas_v3(ineps, N_BT, N_PM, N_GD, N_ID, N_FE)

LARG_V3 = {1: 11, 2: 13, 3: 20, 4: 30, 5: 34, 6: 34, 7: 26, 8: 22, 9: 34, 10: 26,
           11: 26, 12: 22, 13: 13, 14: 9, 15: 13, 16: 48, 17: 28, 18: 22, 19: 12,
           20: 12, 21: 13, 22: 34, 23: 12, 24: 26, 25: 22, 26: 38, 27: 42, 28: 20,
           29: 10, 30: 14, 31: 16, 32: 20, 33: 44, 34: 20, 35: 34, 36: 13, 37: 26,
           38: 36, 39: 9, 40: 16, 41: 16, 42: 20, 43: 40, 44: 26, 45: 18, 46: 26,
           47: 26, 48: 9}
WRAP_V3 = {5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 22, 24, 25, 26, 27, 28, 33, 34, 35,
           37, 38, 42, 43, 44, 45, 46, 47}
CENT_V3 = {1, 2, 13, 14, 15, 19, 20, 21, 23, 29, 30, 31, 32, 36, 39, 40, 41, 48}
EST_V3  = [12, 15, 17, 18, 19, 20, 25, 28, 31, 36, 37, 40, 41, 42, 45, 46, 47, 48]

ws3 = monta_reord("Reordenamento 2027 V3", reord.H_V3, reord.B_V3, frm_v3, val_v3,
                  LARG_V3, WRAP_V3, CENT_V3, EST_V3)
dropdown(ws3, 12, len(ineps), formula="'Validações'!$A$2:$A$12",
         titulo="Fundamental 2027", msg="Etapas do Fundamental que a escola mantém em 2027.")
dropdown(ws3, 28, len(ineps), formula="'Validações'!$H$2:$H$8",
         titulo="Validação da fusão", msg="Confirme, recuse ou adie a fusão sugerida.")
dropdown(ws3, 40, len(ineps), itens=["SIM", "NÃO", "EM ANÁLISE"],
         titulo="Manter o anexo?")
dropdown(ws3, 41, len(ineps), itens=["SIM", "NÃO", "EM ANÁLISE"],
         titulo="Encerrar o anexo?")
dropdown(ws3, 42, len(ineps), formula="'Validações'!$J$2:$J$8",
         titulo="Remanejar a oferta do anexo?")
dropdown(ws3, 45, len(ineps), formula="'Validações'!$I$2:$I$8",
         titulo="Reordenamento 2027")

# ==================================================== BASES DE APOIO
# ---- Base Tratada
BT_BLOCOS = (["id"] * 2 + ["hoje"] + ["em"] * 3 + ["eja"] * 2 + ["fund"] +
             ["hoje"] + ["fusao"] + ["sala"] * 2 + ["saida"] * 2 + ["eja"] +
             ["fund"] * 2 + ["uetep"] * 3 + ["eja"] * 3 + ["fusao"] +
             ["anexo"] * 2 + ["sala"] + ["id"] * 2)
escreve(wb.create_sheet("Base Tratada"), H_BASE_TRATADA, BT_BLOCOS, bt_linhas,
        larguras={1: 11, 2: 30, 3: 48, 4: 9, 5: 9, 6: 9, 7: 44, 8: 20, 9: 32,
                  10: 26, 11: 42, 12: 12, 13: 11, 14: 34, 15: 36, 16: 22, 17: 12,
                  18: 13, 19: 14, 20: 12, 21: 40, 22: 34, 23: 13, 24: 14, 25: 46,
                  26: 36, 27: 9, 28: 14, 29: 20, 30: 13},
        wrap={3, 7, 8, 9, 10, 11, 14, 15, 16, 21, 22, 25, 26},
        centro={1, 4, 5, 6, 12, 13, 17, 18, 19, 20, 23, 24, 27, 28})

# ---- Base Anexos
BA_BLOCOS = (["id"] * 4 + ["anexo"] * 2 + ["em"] * 4 + ["fund"] + ["eja"] * 2 +
             ["hoje"] + ["fusao"] + ["saida"] * 3 + ["sala"] + ["estrela"] * 5)
ws = wb.create_sheet("Base Anexos")
escreve(ws, H_BASE_ANEXOS, BA_BLOCOS, ba_linhas,
        larguras={1: 11, 2: 13, 3: 20, 4: 30, 5: 34, 6: 24, 7: 44, 8: 9, 9: 9,
                  10: 9, 11: 32, 12: 42, 13: 22, 14: 26, 15: 40, 16: 12, 17: 12,
                  18: 34, 19: 14, 20: 16, 21: 16, 22: 20, 23: 30, 24: 34},
        wrap={5, 7, 11, 12, 14, 15, 18, 23, 24},
        centro={1, 8, 9, 10, 16, 17, 19, 20, 21, 22})
dropdown(ws, 20, len(ba_linhas), itens=["SIM", "NÃO", "EM ANÁLISE"],
         titulo="Manter o anexo?", msg="O anexo continua funcionando em 2027?")
dropdown(ws, 21, len(ba_linhas), itens=["SIM", "NÃO", "EM ANÁLISE"],
         titulo="Encerrar o anexo?", msg="O anexo será encerrado em 2027?")
dropdown(ws, 22, len(ba_linhas), formula="'Validações'!$J$2:$J$8",
         titulo="Remanejar a oferta?", msg="Para onde vai a oferta que hoje está no anexo.")

# ---- Panorama Municipal (com os contadores vivos)
PM_BLOCOS = ["id"] + ["fund"] * 7 + ["id"] + ["em"] * 2 + ["uetep"] * 3
ws = wb.create_sheet("Panorama Municipal")
cabecalho(ws, H_PANORAMA, PM_BLOCOS, 62)
for r, L in enumerate(pan_linhas, start=2):
    for c, v in enumerate(L, start=1):
        ws.cell(r, c, v)
    n2, n3 = len(ineps) + 1, len(ineps) + 1
    ws.cell(r, 10, "=IF($A{r}=\"\",\"\",SUMIF('Reordenamento 2027 V2'!$C$2:$C${n},$A{r},"
                   "'Reordenamento 2027 V2'!$M$2:$M${n}))".format(r=r, n=n2))
    ws.cell(r, 11, "=IF($A{r}=\"\",\"\",$H{r}-$J{r})".format(r=r))
    ws.cell(r, 12, "=IF($A{r}=\"\",\"\",SUMIF('Reordenamento 2027 V3'!$C$2:$C${n},$A{r},"
                   "'Reordenamento 2027 V3'!$O$2:$O${n}))".format(r=r, n=n3))
    ws.cell(r, 13, "=IF($A{r}=\"\",\"\",$H{r}-$L{r})".format(r=r))
    ws.cell(r, 14, "=IF($A{r}=\"\",\"\",IF($H{r}=0,\"\",$L{r}/$H{r}))".format(r=r))
    ws.cell(r, 14).number_format = "0%"
corpo(ws, len(pan_linhas), 14,
      larguras={1: 26, 2: 16, 3: 15, 4: 16, 5: 14, 6: 16, 7: 18, 8: 18, 9: 16,
                10: 16, 11: 13, 12: 16, 13: 13, 14: 13},
      wrap=set(), centro=set(range(2, 15)))

# ---- Fusão Turmas (a aba que já existia — fusão turma a turma)
fus_src = d["SEDUC:Fusão Turmas"]
ws = wb.create_sheet("Fusão Turmas")
H_FT = [texto_(v) for v in fus_src[0]]
B_FT = ["estrela"] + ["id"] * 3 + ["em"] * 5 + ["fusao"] * 4 + ["saida"] * 5
escreve(ws, H_FT, B_FT[:len(H_FT)],
        [[int(numero_(v)) if j in (9, 13) and v not in (None, "") else v
          for j, v in enumerate(r)] for r in fus_src[1:]],
        larguras={1: 10, 2: 22, 3: 12, 4: 20, 5: 46, 6: 12, 7: 20, 8: 12, 9: 16,
                  10: 11, 11: 30, 12: 34, 13: 12, 14: 11, 15: 30, 16: 34, 17: 14, 18: 14},
        wrap={5, 11, 12, 15, 16}, centro={1, 6, 8, 10, 13, 14, 17, 18})
dropdown(ws, 1, len(fus_src) - 1, itens=["SIM", "NÃO", "EM ANÁLISE"])

# ---- Fusão Entre Escolas (NOVA)
ws = wb.create_sheet("Fusão Entre Escolas")
B_FE = (["estrela"] + ["id"] * 2 + ["em"] * 2 + ["id"] * 2 + ["em"] * 2 +
        ["fund"] * 3 + ["fund"] + ["id"] * 2 + ["em"] * 2 + ["fund"] + ["sala"] +
        ["fusao"] * 4 + ["id"] + ["estrela"])
for k in range(len(fe_linhas)):
    fe_linhas[k][12] = ('=IFERROR(IF(N(VLOOKUP($B{r},\'IDEB Municípios\'!$A$2:$F${n},2,FALSE))=0,'
                        '"— sem IDEB",VLOOKUP($B{r},\'IDEB Municípios\'!$A$2:$F${n},2,FALSE)),"— sem IDEB")'
                        ).format(r=k + 2, n=N_ID + 1)
escreve(ws, H_FUSAO_ESCOLAS, B_FE, fe_linhas,
        larguras={1: 10, 2: 22, 3: 14, 4: 46, 5: 11, 6: 11, 7: 32, 8: 11, 9: 13,
                  10: 15, 11: 13, 12: 15, 13: 12, 14: 11, 15: 32, 16: 12, 17: 14,
                  18: 15, 19: 11, 20: 15, 21: 14, 22: 14, 23: 15, 24: 16, 25: 34},
        wrap={4, 7, 15, 25}, centro=set(range(5, 25)) - {7, 15})
dropdown(ws, 1, len(fe_linhas), itens=["SIM", "NÃO", "EM ANÁLISE"],
         titulo="Aplicar a fusão?", msg="Marque SIM para levar esta fusão ao relatório.")

# ---- Base UETEP (NOVA)
ws = wb.create_sheet("Base UETEP")
B_UE = ["id"] * 4 + ["em"] * 3 + ["hoje"] + ["uetep"] * 3 + ["sala"]
escreve(ws, H_UETEP, B_UE[:len(H_UETEP)], uetep_linhas,
        larguras={1: 11, 2: 14, 3: 22, 4: 32, 5: 16, 6: 46, 7: 34, 8: 12,
                  9: 16, 10: 13, 11: 11, 12: 16},
        wrap={6, 7}, centro={1, 8, 9, 10, 11, 12})

# ---- Base UEJA (NOVA)
ws = wb.create_sheet("Base UEJA")
B_UJ = ["id"] * 4 + ["anexo"] * 2 + ["eja"] * 3 + ["anexo"] + ["estrela"]
escreve(ws, H_UEJA, B_UJ[:len(H_UEJA)], ueja_linhas,
        larguras={1: 11, 2: 14, 3: 22, 4: 32, 5: 26, 6: 36, 7: 10, 8: 12,
                  9: 52, 10: 16, 11: 30},
        wrap={5, 6, 9, 11}, centro={1, 7, 8, 10})
dropdown(ws, 11, len(ueja_linhas), formula="'Validações'!$J$2:$J$8",
         titulo="Decisão 2027", msg="O que fazer com esta oferta de EJA em 2027.")

# ---- IDEB Municípios (NOVA — para preencher)
ws = wb.create_sheet("IDEB Municípios")
escreve(ws, H_IDEB, ["id"] + ["fund"] * 3 + ["id"] * 2,
        [[m, "", "", "", "", ""] for m in municipios],
        larguras={1: 26, 2: 22, 3: 22, 4: 18, 5: 16, 6: 40},
        wrap={6}, centro={2, 3, 4, 5})

# ---- Cursos (lista oficial 2027)
cur = d["REORD:Cursos"]
ws = wb.create_sheet("Cursos")
escreve(ws, [texto_(v) for v in cur[1]], ["id"] * 3 + ["em"] * 3 + ["saida"],
        [list(r) for r in cur[2:] if texto_(r[0])],
        larguras={1: 8, 2: 52, 3: 26, 4: 12, 5: 12, 6: 10, 7: 20},
        wrap={2}, centro={1, 4, 5, 6})

# ---- Validações (listas dos menus)
FUNDAMENTAL = ["EF Inicial - 1º ano", "EF Inicial - 2º ano", "EF Inicial - 3º ano",
               "EF Inicial - 4º ano", "EF Inicial - 5º ano", "EF Final - 6º ano",
               "EF Final - 7º ano", "EF Final - 8º ano", "EF Final - 9º ano",
               "Não oferta Fundamental", "Encerrar oferta de Fundamental"]
EJA1 = ["Módulo I — 1º ano (Alfabetiza PI) · Início",
        "Módulo II — 2º ano · Continuidade", "Módulo III — 3º ano · Continuidade",
        "Módulo IV — 4º ano · Continuidade", "Módulo V — 5º ano · Continuidade",
        "Módulo VI — 5º ano · Continuidade"]
EJA2 = ["Módulo VII — 6º ano · Continuidade", "Módulo VIII — 7º ano · Continuidade",
        "Módulo IX — 8º ano · Continuidade", "Módulo X — 9º ano · Continuidade"]
EJA3T = ["Módulo I — 1ª série · Início", "Módulo II — 1ª série · Início",
         "Módulo III — 2ª série · Continuidade", "Módulo IV — 2ª série · Continuidade",
         "Módulo V — 3ª série · Finalização"]
EJA3F = ["Módulo I — 1ª série · Início", "Módulo II — 2ª série · Continuidade",
         "Módulo III — 3ª série · Continuidade", "Módulo IV — 3ª série · Finalização"]
TURNOS = ["Integral", "Manhã", "Tarde", "Noite"]
MOVEJA = ["Em continuidade", "Finalizando", "A iniciar (entrada nova)"]
VAL_FUSAO = ["Fusão confirmada", "Fusão recusada", "Fusão em análise",
             "Fusão apenas na própria escola", "Fusão entre escolas do município",
             "Depende de transporte escolar", "Sem fusão possível"]
DECISAO = ["Manter", "Ampliar oferta", "Reduzir oferta", "Fundir com outra escola",
           "Receber turmas de outra escola", "Mudar de turno", "Encerrar oferta"]
DEST_ANEXO = ["Manter no anexo", "Trazer para o prédio matriz",
              "Transferir para outra escola estadual", "Transferir para a rede municipal",
              "Encerrar a oferta", "Em análise"]

cols_val = [("FUNDAMENTAL", FUNDAMENTAL, "fund"),
            ("EJA · Seg I\nEF Inicial", EJA1, "eja"),
            ("EJA · Seg II\nEF Final", EJA2, "eja"),
            ("EJA · Seg III\nEM Técnico", EJA3T, "eja"),
            ("EJA · Seg III\nEM FIC", EJA3F, "eja"),
            ("TURNOS", TURNOS, "id"),
            ("MOVIMENTO\nEJA", MOVEJA, "fusao"),
            ("VALIDAÇÃO\nDA FUSÃO", VAL_FUSAO, "fusao"),
            ("REORDENAMENTO\n2027", DECISAO, "decisao"),
            ("DESTINO DA OFERTA\nDO ANEXO", DEST_ANEXO, "anexo")]
maxl = max(len(c[1]) for c in cols_val)
linhas_val = [[c[1][r] if r < len(c[1]) else "" for c in cols_val] for r in range(maxl)]
escreve(wb.create_sheet("Validações"), [c[0] for c in cols_val],
        [c[2] for c in cols_val], linhas_val,
        larguras={i: 30 for i in range(1, 11)}, wrap=set(), centro=set(), altura=46)

# ---- bases brutas
def bruta(nome, linhas, larguras=None, wrap=None, inteiros=()):
    ws = wb.create_sheet(nome)
    h = [texto_(v) for v in linhas[0]]
    corpo_ = []
    for r in linhas[1:]:
        L = list(r)
        for j in inteiros:
            if j - 1 < len(L) and L[j - 1] not in (None, ""):
                try: L[j - 1] = int(float(L[j - 1]))
                except (TypeError, ValueError): pass
        corpo_.append(L)
    escreve(ws, h, ["id"] * len(h), corpo_,
            larguras=larguras or {}, wrap=wrap or set(), centro=set(), altura=34)

bruta("Turmas", d["SEDUC:Turmas"] if not LIMITE else d["SEDUC:Turmas"][:60],
      {1: 11, 2: 32, 3: 30, 4: 44, 5: 16, 6: 16, 7: 11, 8: 11, 9: 12, 10: 9, 11: 12},
      {4}, inteiros=(1,))
bruta("Base GDI", d["SEDUC:Base GDI"] if not LIMITE else [d["SEDUC:Base GDI"][0]]+[r for r in d["SEDUC:Base GDI"][1:] if inep_(r[0]) in set(ordem_m)],
      {1: 11, 3: 16, 4: 22, 5: 32, 6: 9, 14: 40, 15: 40, 18: 30}, inteiros=(1,))
bruta("Matriculas por etapa", d["REORD:Matriculas por etapa"] if not LIMITE else d["REORD:Matriculas por etapa"][:60],
      {3: 22, 6: 32, 10: 16, 11: 40, 13: 34, 15: 46}, inteiros=(5,))

# ==================================================================== LEIA-ME
ws = wb["LEIA-ME"]
wb.move_sheet("LEIA-ME", offset=-wb.sheetnames.index("LEIA-ME"))
TEXTO = [
 ("T", "REORDENAMENTO 2027 — SEDUC-PI / SUPEX / UGERF / GDI"),
 ("S", "Três versões da aba Reordenamento 2027 no mesmo arquivo, para comparar antes de decidir."),
 ("", ""),
 ("H", "AS TRÊS VERSÕES"),
 ("L", "Reordenamento 2027  —  V1. Retrato exato da aba que existe hoje na Proposta. Nada foi recalculado: é o ponto de partida para comparação."),
 ("L", "Reordenamento 2027 V2  —  mesma estrutura e mesmas 38 colunas da V1, com o que estava sem funcionar já funcionando."),
 ("L", "Reordenamento 2027 V3  —  a V2 mais os pedidos novos: anexos, fusão entre escolas, UETEP, UEJA, fundamental e 2ª/3ª série editáveis."),
 ("", ""),
 ("H", "O QUE ESTAVA QUEBRADO NA V1 E FOI CORRIGIDO NA V2"),
 ("L", "Ainda faltam no município — devolvia \"—\" em todas as 576 linhas. O município vinha acentuado da Base GDI e o Panorama Municipal guarda o nome sem acento, então o PROCV nunca achava. Agora as duas pontas usam o nome sem acento e a coluna responde ao que for digitado na 1ª série."),
 ("L", "Panorama Municipal — as colunas Turmas decididas / Saldo / Cobertura estavam com #ERROR!. Usavam coluna inteira ($C:$C e $M:$M), o que fecha um ciclo de dependência no Google Sheets. Agora o intervalo é limitado às linhas que existem."),
 ("L", "Salas necessárias 2027 — estava preenchida em 1 das 576 linhas. Como Situação da sala e Resumo 2027 dependem dela, as duas saíam vazias (\"Necessidade:  salas |\"). Agora todas vêm calculadas pelo motor e continuam editáveis."),
 ("L", "Possíveis fusões de turmas — devolvia \"—\" em toda a coluna. Voltou a ser calculada, e a coluna ao lado conta quantas turmas saem e quantas chegam segundo a aba Fusão Turmas."),
 ("L", "IDEB, ANEXOS e CURSOS — colunas existiam no cabeçalho e nunca foram preenchidas. Agora saem das abas IDEB Municípios, Base Anexos e Base UETEP."),
 ("", ""),
 ("H", "O QUE A V3 ACRESCENTA"),
 ("L", "ANEXOS — as três perguntas (manter / encerrar / remanejar a oferta) ficam na V3 por escola e na aba Base Anexos uma linha por anexo, que é onde a decisão realmente cabe: 96 escolas têm anexo e algumas têm até 5."),
 ("L", "FUSÃO ENTRE ESCOLAS — aba nova. Cruza o mesmo curso de 1ª série entre escolas do mesmo município, mostra as matrículas dos dois lados, as turmas de fundamental, o 9º ano e o IDEB do município, e diz quantas turmas dá para liberar."),
 ("L", "UETEP — a base Matriculas por etapa entrou no arquivo. Turma sem enturmados e com pré-matrícula é oferta nova: as matrículas novas viram turmas previstas e entram na conta de salas de 2027."),
 ("L", "UEJA — aba nova. Separa a oferta de EJA por onde ela acontece: prédio matriz, anexo, sala externa, comunidade, sala cedida em outra escola, privação de liberdade e socioeducativo."),
 ("L", "FUNDAMENTAL — o Panorama Municipal agora soma o 9º ano da própria rede estadual com o que vem das outras redes do município, e recalcula as turmas necessárias a partir dessa demanda total."),
 ("L", "GERAL — 2ª e 3ª série viraram colunas ★ editáveis. Cada turma que você muda mexe na hora na coluna Salas necessárias 2027 (calculado)."),
 ("", ""),
 ("H", "COMO LER AS CORES"),
 ("L", "Amarelo com ★ — é seu, preencha à mão. Todo o resto é calculado e volta a ser reescrito quando o script roda."),
 ("L", "Cinza escuro = identificação · azul = retrato de 2026 · ciano = fundamental e 9º ano · verde = ensino médio · verde-azulado = UETEP · roxo = fusões · âmbar = salas · marrom = EJA · violeta = anexos."),
 ("", ""),
 ("H", "ABAS DE APOIO"),
 ("L", "Base Tratada — uma linha por prédio matriz. É a fonte de quase tudo que as três versões mostram."),
 ("L", "Base Anexos — uma linha por anexo ou sala externa, com as colunas de decisão."),
 ("L", "Panorama Municipal — demanda de 1ª série por município e o contador de turmas que ainda faltam distribuir."),
 ("L", "Fusão Turmas — fusão turma a turma, a aba que já existia. Fusão Entre Escolas — a sugestão nova, por município."),
 ("L", "Base UETEP · Base UEJA · IDEB Municípios · Cursos · Validações — bases de apoio dos menus e dos cálculos."),
 ("L", "Turmas · Base GDI · Matriculas por etapa — dados brutos, do jeito que chegaram."),
 ("", ""),
 ("H", "O QUE PRECISA SER PREENCHIDO À MÃO"),
 ("L", "IDEB Municípios — os 224 municípios já estão listados, as notas estão em branco. Enquanto não forem preenchidas, a coluna IDEB mostra \"— sem IDEB cadastrado\"."),
 ("", ""),
 ("H", "O SCRIPT"),
 ("L", "O arquivo Reordenamento_2027.gs traz o mesmo cálculo em Apps Script. Cole em Extensões → Apps Script, salve, recarregue a planilha e use o menu 📊 GDI / UGERF."),
 ("L", "Atualizar tudo — refaz Base Tratada, Base Anexos, Base UETEP, Base UEJA, Fusão Entre Escolas e Panorama, e reescreve as colunas calculadas das três versões sem tocar nas colunas ★."),
]
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 150
r = 1
for tipo, txt in TEXTO:
    c = ws.cell(r, 2, txt)
    if tipo == "T":
        c.font = Font(name=ARIAL, size=16, bold=True, color="111827"); ws.row_dimensions[r].height = 26
    elif tipo == "S":
        c.font = Font(name=ARIAL, size=11, color="4B5563"); ws.row_dimensions[r].height = 20
    elif tipo == "H":
        c.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="374151")
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
    else:
        c.font = Font(name=ARIAL, size=10, color="1F2937")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 14 * (1 + len(txt) // 118))
    r += 1
ws.sheet_view.showGridLines = False

wb.save(SAIDA)
print("gerado:", SAIDA)
print("abas:", wb.sheetnames)
