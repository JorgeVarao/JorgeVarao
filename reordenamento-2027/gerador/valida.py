# -*- coding: utf-8 -*-
"""Confere estaticamente todas as fórmulas do arquivo gerado."""
import re, sys
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

ARQ = "Reordenamento_2027_v1_v2_v3.xlsx"
wb = openpyxl.load_workbook(ARQ)
dims = {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}

RE_REF   = re.compile(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d*)(?::\$?([A-Z]{1,3})\$?(\d*))?")
RE_VLOOK = re.compile(r"VLOOKUP\(([^,]+),'([^']+)'!\$([A-Z]{1,3})\$(\d+):\$([A-Z]{1,3})\$(\d+),(\d+),")
RE_LOCAL = re.compile(r"(?<![A-Z0-9'!$])\$?([A-Z]{1,3})\$?(\d+)(?![0-9(])")

problemas, checados = [], 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="): continue
            checados += 1
            # 1. abas referenciadas existem
            for m in RE_REF.finditer(v):
                aba = m.group(1)
                if aba not in dims:
                    problemas.append((ws.title, cell.coordinate, "aba inexistente: " + aba))
            # 2. VLOOKUP: índice cabe no intervalo e o intervalo cabe na aba
            for m in RE_VLOOK.finditer(v):
                _, aba, c1, r1, c2, r2, idx = m.groups()
                if aba not in dims: continue
                larg = column_index_from_string(c2) - column_index_from_string(c1) + 1
                if int(idx) > larg:
                    problemas.append((ws.title, cell.coordinate,
                                      "VLOOKUP índice %s > largura %d em %s" % (idx, larg, aba)))
                if int(r2) > dims[aba][0]:
                    problemas.append((ws.title, cell.coordinate,
                                      "intervalo passa da última linha de %s (%d > %d)"
                                      % (aba, int(r2), dims[aba][0])))
                if column_index_from_string(c2) > dims[aba][1]:
                    problemas.append((ws.title, cell.coordinate,
                                      "intervalo passa da última coluna de %s" % aba))
            # 3. referências locais apontam para a própria linha
            semaba = RE_REF.sub("", v)
            for m in RE_LOCAL.finditer(semaba):
                col, lin = m.group(1), int(m.group(2))
                if lin != cell.row and lin not in (1, 2):
                    problemas.append((ws.title, cell.coordinate,
                                      "referência local fora da linha: %s%d" % (col, lin)))

print("fórmulas conferidas:", checados)
if problemas:
    print("PROBLEMAS:", len(problemas))
    vistos = set()
    for p in problemas:
        k = (p[0], p[2])
        if k in vistos: continue
        vistos.add(k); print("  ", p)
else:
    print("nenhum problema estrutural encontrado")
