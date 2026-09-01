# -*- coding: utf-8 -*-
"""Retrato do trabalho manual existente, antes de qualquer alteração."""
import openpyxl, json
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter, column_index_from_string as ci

ATUAL = "/root/.claude/uploads/41fbd151-f5e2-530d-88f9-9b26c68b08f9/51670d45-Proposta_Reordenamento_2027_SEDUC__PI_1.xlsx"
MEU   = "Reordenamento_2027_v1_v2_v3.xlsx"

def chave(v):
    try: return str(int(float(v)))
    except (TypeError, ValueError): return str(v).strip()

def formula(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v

wv = openpyxl.load_workbook(ATUAL, data_only=True)
wf = openpyxl.load_workbook(ATUAL, data_only=False)
wm = openpyxl.load_workbook(MEU, data_only=True)

# ---- 1. toda célula digitada (não-fórmula, não-vazia) das abas de decisão
digitado = {}
for aba in ["Reordenamento 2027", "Reordenamento 2027 V2", "Reordenamento 2027 V3",
            "Base Anexos (1)", "Base UEJA", "Fusão Entre Escolas", "Fusão Turmas"]:
    ws, wsf = wv[aba], wf[aba]
    heads = {c: str(ws.cell(1, c).value or "").replace("\n", " ") for c in range(1, ws.max_column + 1)}
    cels = {}
    for r in range(2, ws.max_row + 1):
        k = chave(ws.cell(r, 1).value)
        if not k or k == "None": continue
        for c in range(1, ws.max_column + 1):
            f = formula(wsf.cell(r, c))
            if isinstance(f, str) and f.startswith("="): continue
            v = ws.cell(r, c).value
            if v in (None, "") or str(v).strip() == "": continue
            cels.setdefault(c, {})[k] = v
    digitado[aba] = {"heads": heads, "cels": cels}

# ---- 2. o que diverge do que eu entreguei = decisão humana de verdade
print("=" * 84)
print("COLUNAS COM CONTEÚDO DIGITADO  —  o que é semente minha e o que é decisão de vocês")
print("=" * 84)
reais = {}
for aba, d in digitado.items():
    if aba not in wm.sheetnames: continue
    wmv = wm[aba]
    meu = {}
    for r in range(2, wmv.max_row + 1):
        k = chave(wmv.cell(r, 1).value)
        if k and k != "None":
            meu[k] = {c: wmv.cell(r, c).value for c in range(1, wmv.max_column + 1)}
    for c, valores in sorted(d["cels"].items()):
        difs = {}
        for k, v in valores.items():
            antes = meu.get(k, {}).get(c)
            def n(x):
                try: return round(float(x), 4)
                except (TypeError, ValueError): return str(x).strip() if x is not None else None
            if n(antes) != n(v): difs[k] = (antes, v)
        if not valores: continue
        rot = "IGUAL À SEMENTE" if not difs else f"→ {len(difs)} CÉLULA(S) DE VOCÊS"
        print(f"  {aba[-2:]:>3s} {get_column_letter(c):>3s} {d['heads'][c][:40]:42s} {len(valores):4d} preench.  {rot}")
        if difs:
            reais[(aba, c)] = difs
            for k, (a, b) in list(difs.items())[:3]:
                print(f"          INEP {k}: entreguei {a!r} · está {b!r}")

json.dump({f"{a}|{c}": v for (a, c), v in reais.items()}, open("decisoes_reais.json", "w"), default=str)
print(f"\n  células digitadas por vocês (fora da semente): {sum(len(v) for v in reais.values())}")
